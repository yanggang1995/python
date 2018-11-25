# _*_ coding:utf-8 _*_
from xml.etree.ElementTree import *
import openpyxl

SEPARATORS = [" ", "\n", ";", "\t"]
TITLES = ["任务ID", "父任务ID", "写入表", "中间表", "任务名称"]

def readByElementTree(xmlFile, excelFile):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.title = str(excelFile.split("/")[len(excelFile.split("/"))-1].split(".")[0])
    root = parse(xmlFile)
    rows = root.findall("job_def")
    rowCount = 0
    for row in rows:
        rowCount += 1
        resultLine = [row.find("job_id").text, row.find("parent_job_id").text]
        tables = row.find("sql_statement").text.lower().split("insert")
        insertTable = ''
        midTable = ''
        if len(tables) > 1:
            for i in range(1, len(tables)):
                insertSplitList = tables[i].split("}.")
                insertTable += getFirstStrBySeparator(insertSplitList[1], SEPARATORS)
                if i != len(tables) - 1:
                    insertTable += ","
                for j in range(2, len(insertSplitList)):
                    midTable += getFirstStrBySeparator(insertSplitList[j], SEPARATORS)
                    if j != len(insertSplitList) - 1:
                        midTable += ","
        else:
            beforInsert = tables[0].split("}.")
            if midTable != "" and len(beforInsert) > 1:
                midTable += ","
            for i in range(1, len(beforInsert)):
                midTable += getFirstStrBySeparator(beforInsert[i], SEPARATORS)
                if i != len(beforInsert)-1:
                    midTable += ","
        resultLine.append(insertTable)
        resultLine.append(midTable)
        resultLine.append(row.find('job_name').text)
        print(resultLine)
        for i in range(0, len(TITLES)):
            sheet.cell(row=1, column=i+1, value=TITLES[i])
            sheet.column_dimensions[chr(65+i)].width = 30.0
        for colCount in range(0, len(resultLine)):
            sheet.cell(row=rowCount+1, column=colCount+1, value=resultLine[colCount]).number_format = openpyxl.styles.numbers.FORMAT_TEXT

    wb.save(excelFile)


def getFirstStrBySeparator(sqlStr, separators):
    for separator in separators:
        sqlStr = sqlStr.split(separator)[0]
    return sqlStr


def main():
    readByElementTree("../data/xml/nmggd.xml", "../data/excel/nmggd.xlsx")


if __name__ == '__main__':
    main()
