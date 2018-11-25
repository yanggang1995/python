# _*_ coding:UTF-8 _*_

import openpyxl


def readExcel(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.worksheets[0]
    for row in sheet.rows:
        for cell in row :
            print(cell.value)

def writeExcel(file):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "test"
    for i in range(0,4):
        for j in range(0, 4):
            sheet.cell(i+1, j+1).number_format = openpyxl.styles.numbers.FORMAT_TEXT
            sheet.cell(i+1, j+1).value = str(i)+"---"+str(j)
    wb.save(file)
    print("写入成功")

def main():
    writeExcel("../data/excel/test01.xlsx")
    readExcel("../data/excel/test01.xlsx")

if __name__ == '__main__':
    main()
